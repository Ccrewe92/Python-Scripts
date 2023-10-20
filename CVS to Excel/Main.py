import openpyxl
import sys

def get_user_input():
    # Prompt the user for required inputs
    print("This program transfers data from CSV (.csv or .data) to an Excel file.")
    print("Both the input and output files should be in the same directory as this script.\n")

    csv_name = input("Enter the CSV filename (including extension): ")
    sep = input("Enter the CSV delimiter: ")
    excel_name = input("Enter the desired Excel filename (including extension): ")
    sheet_name = input("Enter the desired Excel sheet name: ")
    return csv_name, sep, excel_name, sheet_name

def write_to_excel(csv_name, sep, excel_name, sheet_name):
    """Write data from CSV to the specified Excel sheet."""
    try:
        # Read content from the CSV file
        with open(csv_name, "r", encoding="utf-8") as file:
            content = file.readlines()
            
        # Load or create the Excel workbook and sheet
        wb = openpyxl.load_workbook(excel_name)
        if sheet_name in wb:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(title=sheet_name)
        
        # Populate the Excel sheet with data from CSV
        for row_idx, line in enumerate(content, start=1):
            values = line.strip().split(sep)
            for col_idx, data in enumerate(values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=data)

        # Save the populated Excel file
        wb.save(excel_name)
    except Exception as e:
        print(f"Encountered an error: {e}")
        sys.exit()

def main():
    """Main execution function."""
    csv_name, sep, excel_name, sheet_name = get_user_input()
    write_to_excel(csv_name, sep, excel_name, sheet_name)

if __name__ == "__main__":
    main()